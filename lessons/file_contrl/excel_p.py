#!/usr/bin/python3
# -*- coding: utf-8 -*-
# @Time     : 2021/3/31 21:00
# @Author   : ZhangTao
# @File     : excel_p.py
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

wb = Workbook()
dest_filename = 'empty_book.xlsx'

ws1 = wb.active # 调用正在运行的工作表
ws1.title = "range names"   # 标题

for row in range(1, 40):
    ws1.append(range(600))

ws2 = wb.create_sheet(title="Pi")

ws2['F5'] = 3.14

ws3 = wb.create_sheet(title="Data")
for row in range(10, 20):
    for col in range(27, 54):
        _ = ws3.cell(column=col, row=row, value="{0}".format(get_column_letter(col)))
print(ws3['AA10'].value)

ws4 = wb.create_sheet(title="MySheet")
for i in range(1, 30):
    for j in range(3, 20):
        ws4.cell(column=j, row=i).value='test'

wb.save(filename = dest_filename)