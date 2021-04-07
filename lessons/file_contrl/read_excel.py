#!/usr/bin/python3
# -*- coding: utf-8 -*-
# @Time     : 2021/3/31 21:09
# @Author   : ZhangTao
# @File     : read_excel.py
from openpyxl import load_workbook
wb = load_workbook(filename ='empty_book.xlsx')
sheet_ranges = wb['range names']
print(sheet_ranges['D18'].value)

# 打印第一行前30列数据
for i in range(1, 31):
    print(sheet_ranges.cell(column=i, row=1).value)